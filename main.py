import time
import requests
import openpyxl

from openpyxl.drawing.image import Image
from io import BytesIO
from fake_useragent import UserAgent

UsAgent = UserAgent().random

headers = {
    'accept': '*/*',
    'user-agent': UsAgent,
    'x-app-version': '0.1.0',
    'x-client-name': 'magnit',
    'x-device-id': '123456789',
    'x-device-platform': 'Web',
    'x-device-tag': 'disabled',
    'x-platform-version': 'window.navigator.userAgent',
}

wb = openpyxl.Workbook()
ws = wb.active

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 75
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15

address = input("Введите адрес магазина: ")
carry = {"query": address, "storeTypes": [1, 2, 6, 5], "limit": 2000}
shop = requests.post(url='https://magnit.ru/webgate/v1/store-search/address', headers=headers, json=carry)
shop_info = shop.json()

print(f"(!) Выбранный магазин: [{shop_info['stores'][0]['code']}] «{shop_info['stores'][0]['name']}» ({shop_info['stores'][0]['address']})")
# target = input("(?) Что вы хотите найти: ")
input("Для начала нажмите ENTER...")

offset = 1
limit = 50

ws[f'A{offset}'] = "Изображение"
ws[f'B{offset}'] = "Название"
ws[f'C{offset}'] = "Категория"
ws[f'D{offset}'] = "Начало"
ws[f'E{offset}'] = "Окончание"
ws[f'F{offset}'] = "Цена"

while True:

    params = {
        "offset": offset,
        "limit": limit,
        "storeCode": shop_info['stores'][0]['code'],
        "adult": True
    }

    goods = requests.get(url='https://magnit.ru/webgate/v1/promotions', headers=headers, params=params)
    goods_info = goods.json()['data']

    if not goods_info:
        print("Товаров нет!")
        wb.save(f'result{offset}.xlsx')
        break

    for product in goods_info:
        offset += 1
        print(f"{offset - 1}) {product['name']}")

        image_url = product['image']
        response = requests.get(image_url)
        if response.status_code:
            img_data = BytesIO(response.content)
        else:
            print("Не удалось скачать изображение!")
            continue

        try:
            img = Image(img_data)
            img.width = 100
            img.height = 100
            ws.add_image(img, f'A{offset}')
        except Exception:
            pass

        ws.row_dimensions[offset].height = 75
        ws[f'B{offset}'] = product['name']
        ws[f'C{offset}'] = product['categoryName']
        ws[f'D{offset}'] = product['startDate']
        ws[f'E{offset}'] = product['endDate']

        if 'price' in product.keys():
            ws[f'F{offset}'] = product['price']/100

    print("\nПереход на другую страницу...\n")
    time.sleep(3)
